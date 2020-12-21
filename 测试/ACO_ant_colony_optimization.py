import numpy as np
import random
import openpyxl
import matplotlib.pyplot as plt
import copy as cp

MPS=3
big_m = 1000

class data:
    F=15
    R=2
    workbook = openpyxl.load_workbook('data.xlsx')
    data1 = workbook['Sheet1']
    data2 = workbook['Sheet2']
    data3 = workbook['Sheet3']
    
    def __init__(self):
        E = []
        ETA = []
        L = []
        city_name = []
        Fnum = range(1, data.F + 1)
        FType = []
        for i in range(1, data.F + 1):
            if data.data1.cell(row=i, column=2).value == 'J':
                FType.append(1)
            elif data.data1.cell(row=i, column=2).value == 'B':
                FType.append(2)
            elif data.data1.cell(row=i, column=2).value == 'C':
                FType.append(3)
            elif data.data1.cell(row=i, column=2).value == 'M':
                FType.append(4)
            else:
                FType.append(5)  # 把字符形式J、B、C、M、L转化成1,2,3，4,5
            city_name.append(data.data1.cell(row=i,column=1).value)
            E.append(data.data1.cell(row=i, column=3).value)  # 读取E,ETA,L数据
            ETA.append(data.data1.cell(row=i, column=4).value)
            L.append(data.data1.cell(row=i, column=5).value)          
        SMatrix = np.zeros((5, 5))
        for i in range(1, 6):
            for j in range(1, 6):
                SMatrix[i - 1][j - 1] = data.data2.cell(row=i, column=j).value
         # 构建S矩阵

        S = np.zeros((data.F, data.F))
        for i in range(0, data.F):
            for j in range(0, data.F):
                S[i][j] = SMatrix[FType[i] - 1][FType[j] - 1]
                if i == j:
                    S[i][j] = big_m

        DMatrix = np.zeros((15, 15))
        for i in range(1, 16):
            for j in range(1, 16):
                DMatrix[i - 1][j - 1] = data.data3.cell(row=i, column=j).value
         # D矩阵
         

        D = np.zeros((data.F, data.F))
        for i in range(0, data.F):
            for j in range(0, data.F):
                D[i][j] = DMatrix[FType[i] - 1][FType[j] - 1]
                if i == j:
                    D[i][j] = 10000
                    
        self.E=E
        self.ETA=ETA
        self.L=L
        self.F=data.F
        self.S=S
        self.D=D
        self.R=data.R
        self.city_name=city_name
Data=data()


class Individual:
    '个体类，包括构造个体，计算个体适应度值'

    def __init__(self,Data):
        self.F=Data.F
        self.R=Data.R
        self.L=np.array(Data.L)
        self.ETA=np.array(Data.ETA)
        self.E=np.array(Data.E)
        self.S=Data.S
        self.NXP=np.arange(1,self.F+1)
        self.OXP = np.arange(1, self.F + 1)


    def generate_NXP(self,NXP=[]):
        
        AntCount = 15
        # 城市数量
        city_count = len(Data.city_name)

        alpha = 1  # 信息素重要程度因子
        beta = 2  # 启发函数重要程度因子
        rho = 0.1 #挥发速度
        iter = 0  # 迭代初始值
        MAX_iter = 200  # 最大迭代值
        Q = 1
        # 初始信息素矩阵，全是为1组成的矩阵
        pheromonetable = np.ones((city_count, city_count))

        # 候选集列表,存放30只蚂蚁的路径(一只蚂蚁一个路径),一共就Antcount个路径，一共是蚂蚁数量*31个城市数量
        candidate = np.zeros((AntCount, city_count)).astype(int) 

        # path_best存放的是相应的，每次迭代后的最优路径，每次迭代只有一个值
        path_best = np.zeros((MAX_iter, city_count)) 

        # 存放每次迭代的最优距离
        distance_best = np.zeros( MAX_iter)
        # 倒数矩阵
        etable = 1.0 / Data.D

        if any(NXP):
            self.NXP=NXP
        else:
              
            while iter <  MAX_iter:
            # first：蚂蚁初始点选择
                if AntCount <= city_count:
                #np.random.permutation随机排列一个数组的
                    candidate[:, 0] = np.random.permutation(range(city_count))[:AntCount]
                else:
                    m =AntCount -city_count
                    n =2
                    candidate[:city_count, 0] = np.random.permutation(range(city_count))[:]
                    while m >city_count:
                        candidate[city_count*(n -1):city_count*n, 0] = np.random.permutation(range(city_count))[:]
                        m = m -city_count
                        n = n + 1
                    candidate[city_count*(n-1):AntCount,0] = np.random.permutation(range(city_count))[:m]
                length = np.zeros(AntCount)#每次迭代的N个蚂蚁的距离值

                # second：选择下一个城市选择
                for i in range(AntCount):
                # 移除已经访问的第一个元素
                    unvisit = list(range(city_count))  # 列表形式存储没有访问的城市编号
                    visit = candidate[i, 0]  # 当前所在点,第i个蚂蚁在第一个城市
                    unvisit.remove(visit)  # 在未访问的城市中移除当前开始的点
                    for j in range(1, city_count):#访问剩下的city_count个城市，city_count次访问
                        protrans = np.zeros(len(unvisit))#每次循环都更改当前没有访问的城市的转移概率矩阵1*30,1*29,1*28...
                        # 下一城市的概率函数
                        for k in range(len(unvisit)):
                        # 计算当前城市到剩余城市的（信息素浓度^alpha）*（城市适应度的倒数）^beta
                        # etable[visit][unvisit[k]],(alpha+1)是倒数分之一，pheromonetable[visit][unvisit[k]]是从本城市到k城市的信息素
                            protrans[k] = np.power(pheromonetable[visit][unvisit[k]], alpha) * np.power(
                               etable[visit][unvisit[k]], (alpha + 1))

                        # 累计概率，轮盘赌选择
                        cumsumprobtrans = (protrans / sum(protrans)).cumsum()
                        cumsumprobtrans -= np.random.rand()
                        # 求出离随机数产生最近的索引值
                        k = unvisit[list(cumsumprobtrans > 0).index(True)]
                        # 下一个访问城市的索引值
                        candidate[i, j] = k
                        unvisit.remove(k)
                        length[i] += Data.D[visit][k]
                        visit = k  # 更改出发点，继续选择下一个到达点
                    length[i] += Data.D[visit][candidate[i, 0]]#最后一个城市和第一个城市的距离值也要加进去
                if iter == 0:
                    distance_best[iter] = length.min()
                    path_best[iter] = candidate[length.argmin()].copy()
                else:
                # 如果当前的解没有之前的解好，那么当前最优还是为之前的那个值；并且用前一个路径替换为当前的最优路径
                    if length.min() > distance_best[iter - 1]:
                        distance_best[iter] = distance_best[iter - 1]
                        path_best[iter] = path_best[iter - 1].copy()
                    else:  # 当前解比之前的要好，替换当前解和路径
                        distance_best[iter] = length.min()
                        path_best[iter] = candidate[length.argmin()].copy()

                changepheromonetable = np.zeros((city_count, city_count))
                for i in range(AntCount):
                    for j in range(city_count - 1):
                    # 当前路径比如城市23之间的信息素的增量：1/当前蚂蚁行走的总距离的信息素
                        changepheromonetable[candidate[i, j]][candidate[i][j + 1]] += Q / length[i]
                    #Distance[candidate[i, j]][candidate[i, j + 1]]
                    #最后一个城市和第一个城市的信息素增加量
                        changepheromonetable[candidate[i, j + 1]][candidate[i, 0]] += Q / length[i]
                #信息素更新的公式：
                pheromonetable = (1 - rho) * pheromonetable + changepheromonetable
                iter += 1

            self.NXP=path_best[-1]+1

    def generate_runway(self,runway=[]):
        if any(runway):
            self.runway=runway
        else:
            self.runway = np.arange(0, self.F)
            for i in range(0, self.F):
                self.runway[i] = np.random.randint(0, self.R)  # 生成跑道选择层 跑道序号从0到R-1
    def generate_decision_vars(self):
        self.z=np.zeros((self.F,self.R))          #决策变量z[i][r]
        for i,r in zip(self.OXP,self.runway):
            self.z[i-1][r]=1
            self.gamma=np.zeros((self.F,self.F))      #决策变量gamma
        for i in range(0,self.F):
            for j in range(0,self.F):
                if self.runway[i]==self.runway[j]:
                    self.gamma[i][j]=1
        self.delta=np.zeros((self.F,self.F))      #决策变量delta
        for i in range(0,self.F):
            for j in range(0,self.F):
                if self.NXP[i]<self.NXP[j]:
                    self.delta[i][j]=1
         #STA[j]=max{ETA[j],STA[i]+S[i][j]*gamma[i][j]}
        self.STA=np.zeros(self.F)
        t=0
        while self.NXP[t]!=1:
           t=t+1            #找到第一个降落的航班，让其STA=ETA
        self.STA[t]=self.ETA[t]
        for i in range(2,self.F+1):
            k=0
            while self.NXP[k]!=i:
                k=k+1            #找到第i个降落的航班，t为前一个降落的航班在NXP序列中的索引
            self.STA[k]=max(self.ETA[k],self.STA[t]+self.S[k][t]*self.gamma[k][t])
            t=k
    def cal_fitness(self):
        dl=0                                            #计算适应度函数值
        for i in range(0,self.F):
            if self.STA[i]>self.L[i] or self.STA[i]<self.E[i]:
                dl+=big_m                                        #E<ETA<L的约束条件
            if abs(self.NXP[i]-self.OXP[i])<=MPS:
                dl+=self.STA[i]-self.ETA[i]
            else:                                                      #最大位移约束
                dl+=big_m
        dl = 1/(1+dl)   #最小化问题转化为最大化问题
        return dl





class Population:
    '种群类，包括初始化、交叉、变异操作'
            
    def __init__(self,popsize):
        self.popsize=popsize
        self.pop = []
        for i in range(self.popsize):
            p=Individual(Data)
            p.generate_NXP()
            p.generate_runway()
            p.generate_decision_vars()
            self.pop.append(p)
        # 完全随机生成初始种群

    def parent_selection(self):
        '轮盘赌选择父母'
        sum=0
        p=np.zeros(self.popsize)
        for i in range(0,self.popsize):
            sum+=self.pop[i].cal_fitness()
        F=self.pop[0].cal_fitness()
        k=0
        r=random.uniform(0,sum)
        while F<r:
            k=k+1
            F=F+self.pop[k].cal_fitness()
        return k

    def generate_children(self):
        'PMX交叉方法'
        F=self.pop[0].F
        children=[]
        while len(children)<self.popsize:
            left = np.random.randint(0, F-1)
            right = np.random.randint(left+1, F)  # 随机生成两个交叉点left,right
            k1 = self.parent_selection()
            k2 = self.parent_selection()
            p1 = self.pop[k1].NXP
            p2 = self.pop[k2].NXP  # 轮盘赌选择父母
            child1=cp.deepcopy(p1)
            child2=cp.deepcopy(p2)
            r_a_b=range(left,right+1)
            r_left = np.delete(range(F), r_a_b)
            left_1, left_2 = child1[r_left], child2[r_left]
            middle_1, middle_2 = child1[r_a_b], child2[r_a_b]
            child1[r_a_b], child2[r_a_b] = middle_2, middle_1
            mapping = [[], []]
            for i, j in zip(middle_1, middle_2):        #PMX交叉
                if j in middle_1 and i not in middle_2:
                    index = np.argwhere(middle_1 == j)[0, 0]
                    value = middle_2[index]
                    while True:
                        if value in middle_1:
                            index = np.argwhere(middle_1 == value)[0, 0]
                            value = middle_2[index]
                        else:
                            break
                    mapping[0].append(i)
                    mapping[1].append(value)
                elif i in middle_2:
                    pass
                else:
                    mapping[0].append(i)
                    mapping[1].append(j)
            for i, j in zip(mapping[0], mapping[1]):
                if i in left_1:
                    left_1[np.argwhere(left_1 == i)[0, 0]] = j
                elif i in left_2:
                    left_2[np.argwhere(left_2 == i)[0, 0]] = j
                if j in left_1:
                    left_1[np.argwhere(left_1 == j)[0, 0]] = i
                elif j in left_2:
                    left_2[np.argwhere(left_2 == j)[0, 0]] = i
            child1[r_left], child2[r_left] = left_1, left_2
            theta=0.01 #变异率
            if random.random()<theta:
                a=random.randint(0,F-1)
                b=random.randint(0,F-1)
                while a==b:
                    a = random.randint(0, F - 1)
                    b = random.randint(0, F - 1)
                temp=np.array(child1[a])
                child1[a]=child1[b]
                child1[b]=temp
            if random.random()<theta:
                a=random.randint(0,F-1)
                b=random.randint(0,F-1)
                while a==b:
                    a = random.randint(0, F - 1)
                    b = random.randint(0, F - 1)
                temp = np.array(child2[a])
                child2[a] = child2[b]
                child2[b] = temp
            t1=cp.deepcopy(self.pop[k1])
            t2=cp.deepcopy(self.pop[k2])
            t1.generate_NXP(child1)
            t1.generate_decision_vars()
            t2.generate_NXP(child2)
            t2.generate_decision_vars()
            children.append(t1)
            children.append(t2)
        self.pop=children
        

def algorithm_iter(itertimes):
    popsize=500
    #生成初始种群
    Parents=Population(popsize)
    eq=[]
    eqmax=[]
    for j in range(0, popsize):
        eq.append(Parents.pop[j].cal_fitness())
    eqmax.append(max(eq))
    for i in range(0,itertimes):
        Parents.generate_children()
        eq=[]
        for j in range(0,popsize):
            eq.append(Parents.pop[j].cal_fitness())
        eqmax.append(max(eq))
    plt.plot(eqmax)
    plt.show()
    return Parents

if __name__ == '__main__':

    p=algorithm_iter(200)

