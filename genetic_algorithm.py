import numpy as np
import random
import openpyxl
import matplotlib.pyplot as plt
import copy as cp
MPS=3
big_m = 0xFFFFF

class data:
    F=30
    R=2
    workbook = openpyxl.load_workbook('data.xlsx')
    data1 = workbook['Sheet1']
    data2 = workbook['Sheet2']
    data3 = workbook['Sheet3']
    def __init__(self):
        E = []
        ETA = []
        L = []
        Fnum = range(1, data.F + 1)
        FType = []
        for i in range(1, data.F + 1):
            if data.data1.cell(row=i, column=2).value == 'H':
                FType.append(1)
            elif data.data1.cell(row=i, column=2).value == 'M':
                FType.append(2)
            else:
                FType.append(3)  # 把字符形式H、M、L转化成1,2,3
            E.append(data.data1.cell(row=i, column=3).value)  # 读取E,ETA,L数据
            ETA.append(data.data1.cell(row=i, column=4).value)
            L.append(data.data1.cell(row=i, column=5).value)
        SMatrix = np.zeros((3, 3))
        for i in range(1, 4):
            for j in range(1, 4):
                SMatrix[i - 1][j - 1] = data.data2.cell(row=i, column=j).value
        sMatrix = np.zeros((3, 3))
        for i in range(1, 4):
            for j in range(1, 4):
                sMatrix[i - 1][j - 1] = data.data3.cell(row=i, column=j).value  # 构建S,s矩阵

        S = np.zeros((data.F, data.F))
        s = np.zeros((data.F, data.F))
        for i in range(0, data.F):
            for j in range(0, data.F):
                S[i][j] = SMatrix[FType[i] - 1][FType[j] - 1]
                s[i][j] = sMatrix[FType[i] - 1][FType[j] - 1]
                if i == j:
                    s[i][j] = big_m
                    S[i][j] = big_m
        self.E=E
        self.ETA=ETA
        self.L=L
        self.F=data.F
        self.S=S
        self.s=s
        self.R=data.R

Data=data()


class Individual:
    '个体类，包括构造个体，计算个体适应度值'

    def __init__(self,Data):
        self.F=Data.F
        self.R=Data.R
        self.L=np.array(Data.L)
        self.ETA=np.array(Data.ETA)
        self.E=np.array(Data.E)
        self.S=Data.S
        self.s=Data.s
        self.NXP=np.arange(1,self.F+1)
        self.OXP = np.arange(1, self.F + 1)
    def generate_NXP(self,NXP=[]):
        if any(NXP):
            self.NXP=NXP
        else:
            self.NXP = np.arange(1, self.F + 1)
            np.random.shuffle(self.NXP)  # 生成航班排序队列层 1...R排列
    def generate_runway(self,runway=[]):
        if any(runway):
            self.runway=runway
        else:
            self.runway = np.arange(0, self.F)
            for i in range(0, self.F):
                self.runway[i] = np.random.randint(0, self.R)  # 生成跑道选择层 跑道序号从0到R-1
    def generate_decision_vars(self):
        self.z=np.zeros((self.F,self.R))          #决策变量z[i][r]
        for i,r in zip(self.OXP,self.runway):
            self.z[i-1][r]=1
        self.gamma=np.zeros((self.F,self.F))      #决策变量gamma
        for i in range(0,self.F):
            for j in range(0,self.F):
                if self.runway[i]==self.runway[j]:
                    self.gamma[i][j]=1
        self.delta=np.zeros((self.F,self.F))      #决策变量delta
        for i in range(0,self.F):
            for j in range(0,self.F):
                if self.NXP[i]<self.NXP[j]:
                    self.delta[i][j]=1
        #STA[j]=max{ETA[j],STA[i]+S[i][j]*gamma[i][j]+s[i][j](1-delta[i][j])}
        self.STA=np.zeros(self.F)
        t=0
        while self.NXP[t]!=1:
           t=t+1            #找到第一个降落的航班，让其STA=ETA
        self.STA[t]=self.ETA[t]
        for i in range(2,self.F+1):
            k=0
            while self.NXP[k]!=i:
                k=k+1            #找到第i个降落的航班，t为前一个降落的航班在NXP序列中的索引
            self.STA[k]=max(self.ETA[k],self.STA[t]+self.S[k][t]*self.gamma[k][t]+self.s[k][t]*(1-self.delta[k][t]))
            t=k

    def cal_fitness(self):
        dl=0                                            #计算适应度函数值
        for i in range(0,self.F):
            if self.STA[i]>self.L[i] or self.STA[i]<self.E[i]:
                dl+=big_m                                        #E<ETA<L的约束条件
            if abs(self.NXP[i]-self.OXP[i])<=MPS:
                dl+=self.STA[i]-self.ETA[i]
            else:                                                      #最大位移约束
                dl+=big_m
        dl = 1/(1+dl)   #最小化问题转化为最大化问题
        return dl


class Population:
    '种群类，包括初始化、交叉、变异操作'
    def __init__(self,popsize):
        self.popsize=popsize
        self.pop = []
        for i in range(self.popsize):
            p=Individual(Data)
            p.generate_NXP()
            p.generate_runway()
            p.generate_decision_vars()
            self.pop.append(p)
        # 完全随机生成初始种群

    def parent_selection(self):
        '轮盘赌选择父母'
        sum=0
        p=np.zeros(self.popsize)
        for i in range(0,self.popsize):
            sum+=self.pop[i].cal_fitness()
        F=self.pop[0].cal_fitness()
        k=0
        r=random.uniform(0,sum)
        while F<r:
            k=k+1
            F=F+self.pop[k].cal_fitness()
        return k



    def generate_children(self):
        'PMX交叉方法'
        F=self.pop[0].F
        children=[]
        while len(children)<self.popsize:
            left = random.randint(0, F-1)
            right = random.randint(left, F-1)  # 随机生成两个交叉点left,right
            k1 = self.parent_selection()
            k2 = self.parent_selection()
            p1 = self.pop[k1].NXP
            p2 = self.pop[k2].NXP  # 轮盘赌选择父母
            child1=cp.deepcopy(p1)
            child2=cp.deepcopy(p2)
            r_a_b=range(left,right+1)
            r_left = np.delete(range(F), r_a_b)
            left_1, left_2 = child1[r_left], child2[r_left]
            middle_1, middle_2 = child1[r_a_b], child2[r_a_b]
            child1[r_a_b], child2[r_a_b] = middle_2, middle_1
            mapping = [[], []]
            for i, j in zip(middle_1, middle_2):        #PMX交叉
                if j in middle_1 and i not in middle_2:
                    index = np.argwhere(middle_1 == j)[0, 0]
                    value = middle_2[index]
                    while True:
                        if value in middle_1:
                            index = np.argwhere(middle_1 == value)[0, 0]
                            value = middle_2[index]
                        else:
                            break
                    mapping[0].append(i)
                    mapping[1].append(value)
                elif i in middle_2:
                    pass
                else:
                    mapping[0].append(i)
                    mapping[1].append(j)
            for i, j in zip(mapping[0], mapping[1]):
                if i in left_1:
                    left_1[np.argwhere(left_1 == i)[0, 0]] = j
                elif i in left_2:
                    left_2[np.argwhere(left_2 == i)[0, 0]] = j
                if j in left_1:
                    left_1[np.argwhere(left_1 == j)[0, 0]] = i
                elif j in left_2:
                    left_2[np.argwhere(left_2 == j)[0, 0]] = i
            child1[r_left], child2[r_left] = left_1, left_2
            theta=0.01 #变异率
            if random.random()<theta:
                a=random.randint(0,F-1)
                b=random.randint(0,F-1)
                while a==b:
                    a = random.randint(0, F - 1)
                    b = random.randint(0, F - 1)
                temp=np.array(child1[a])
                child1[a]=child1[b]
                child1[b]=temp
            if random.random()<theta:
                a=random.randint(0,F-1)
                b=random.randint(0,F-1)
                while a==b:
                    a = random.randint(0, F - 1)
                    b = random.randint(0, F - 1)
                temp = np.array(child2[a])
                child2[a] = child2[b]
                child2[b] = temp
            t1=cp.deepcopy(self.pop[k1])
            t2=cp.deepcopy(self.pop[k2])
            t1.generate_NXP(child1)
            t1.generate_decision_vars()
            t2.generate_NXP(child2)
            t2.generate_decision_vars()
            children.append(t1)
            children.append(t2)
        self.pop=children


def algorithm_iter(itertimes):
    popsize=500
    #生成初始种群
    Parents=Population(popsize)
    eq=[]
    eqmax=[]
    for j in range(0, popsize):
        eq.append(Parents.pop[j].cal_fitness())
    eqmax.append(max(eq))
    for i in range(0,itertimes):
        Parents.generate_children()
        eq=[]
        for j in range(0,popsize):
            eq.append(Parents.pop[j].cal_fitness())
        eqmax.append(max(eq))
    plt.plot(eqmax)
    plt.show()
    return Parents

if __name__ == '__main__':
    p=algorithm_iter(100)




